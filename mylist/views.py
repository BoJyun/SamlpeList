import datetime
import os
import openpyxl
from rest_framework import status
from mylist.serializers import MylistSerializer
from mylist.models import Mylist, AttachmentFile
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny
from django.utils.http import urlquote
from django.http import FileResponse
from django.conf import settings
from account.verify import JWTAuthentication
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
# Create your views here.


@api_view(['POST'])
@permission_classes([AllowAny])
def getFile(request):
    result = {}
    try:
        serializer = MylistSerializer(data=request.data)
        if serializer.is_valid():
            name = serializer.data['name']
            date = serializer.data['date']
            num = serializer.data['num']
            file = serializer.data['file']
            commit = serializer.data['commit']
            important = serializer.data['important']
            complete = serializer.data['complete']
        else:
            result['msg'] = 'you miss some data...'
            result['success'] = False
            return Response(result, status.HTTP_200_OK)

        fileload = request.FILES.get('product')
        if fileload is None:
            result['msg'] = 'upload file fail please try again.'
            result['success'] = False
            return Response(result, status.HTTP_400_BAD_REQUEST)

        folder = str(date)
        folder_absolute_path = settings.MEDIA_ROOT + \
            r'/'+str(name) + r'/' + folder
        if not os.path.exists(folder_absolute_path):
            build_folder = mkdirs_in_batch(folder_absolute_path)
            if not build_folder:
                result['msg'] = '批量創建路徑{}對應的目錄失敗'.format(folder_absolute_path)
                result['success'] = False
                return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)
        try:
            folder_absolute_path = folder_absolute_path + r'/'+file
            file_handler = open(folder_absolute_path, 'wb')

            for chunk in fileload.chunks():
                file_handler.write(chunk)
            file_handler.close()
            obj = Mylist.objects.create(name=name, date=date, num=num, file=file, commit=commit,
                                        important=important, complete=complete, file_path=folder_absolute_path)

            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)("mylist", {
                "type": "chat_message",
                "message": ["fileChange2", True],
            })

            result['msg'] = "Upload Success"
            result['success'] = True
            return Response(result, status.HTTP_200_OK)

        except Exception as e:
            file_handler.close()
            result['msg'] = '{}'.format(e)
            result['success'] = False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)

    except Exception as e:
        result['msg'] = 'Upload False'
        result['success'] = False
        return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def getCompletedList(request):
    date = datetime.date.today()
    list = Mylist.objects.filter(date=date)
    dataSerializer = MylistSerializer(list, many=True)
    return Response(dataSerializer.data, status=status.HTTP_200_OK,)


@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([JWTAuthentication])
def downloadFile(request):
    result = {}
    try:
        id = request.data.get('id')
        obj = Mylist.objects.filter(id=id)[0]

        if obj:
            file_absoulte_path = obj.file_path
            if os.path.exists(file_absoulte_path):
                try:
                    file_absoulte_path = os.path.normpath(file_absoulte_path)
                    head, file_name = os.path.split(file_absoulte_path)
                    file = open(file_absoulte_path, 'rb')
                    file_response = FileResponse(file)
                    file_response['Content-Type'] = 'application/octet-stream'
                    file_response["Access-Control-Expose-Headers"] = 'Content-Disposition'
                    file_response['Content-Disposition'] = 'attachment;filename={}'.format(
                        urlquote(file_name))
                    return file_response

                except Exception as e:
                    result['msg'] = '{}'.format(e)
                    result['success'] = False
                    return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)

    except Exception as e:
        result['msg'] = '{}'.format(e)
        result['success'] = False
        return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)


@ api_view(['GET'])
@ permission_classes([AllowAny])
def getdata(request):
    mylist = Mylist.objects.all()
    dataSerializer = MylistSerializer(mylist, many=True)
    return Response(dataSerializer.data, status=status.HTTP_200_OK,)


@ api_view(['PATCH'])
@ permission_classes([AllowAny])
@authentication_classes([JWTAuthentication])
def complete(request, id):
    mylist = Mylist.objects.filter(id=id)[0]
    if mylist is not None:
        mylist.complete = True
        mylist.save()
        obj = AttachmentFile.objects.create(file_path=mylist, done_user=request.data.get(
            'done_user'), start_time=request.data.get('startTime'), done_time=request.data.get('endTime'))
        return Response({'Message': 'Update Success'}, status=status.HTTP_200_OK)
    else:

        return Response({'Message': 'Invalid data...'}, status=status.HTTP_400_BAD_REQUEST)


@ api_view(['DELETE'])
@ permission_classes([AllowAny])
def quit(request, id):
    try:
        mylistID = Mylist.objects.filter(id=id)
        mylist = mylistID[0]
        mylist.delete()
        return Response({'Message': 'Delete Success'}, status=status.HTTP_200_OK)

    except Exception as e:
        print("Errors: {}".format(e))
        text = "Errors: {}".format(e)
        return Response({'Message': text}, status=status.HTTP_400_BAD_REQUEST)


@ api_view(['PATCH'])
@ permission_classes([AllowAny])
@authentication_classes([JWTAuthentication])
def update(request, id):
    try:
        mylistID = Mylist.objects.filter(id=id)
        mylist = mylistID[0]
        serializer = MylistSerializer(mylist, data=request.data, partial=True)

        if serializer.is_valid():

            serializer.save()

            return Response({'Message': 'Update Success'}, status=status.HTTP_200_OK)
        else:
            print(serializer.errors)
            return Response({'Message': 'Invalid data...'}, status=status.HTTP_200_OK)

    except Exception as e:
        print("Errors: {}".format(e))
        text = "Errors: {}".format(e)
        return Response({'Message': text}, status=status.HTTP_400_BAD_REQUEST)


def mkdirs_in_batch(path):
    try:
        path = os.path.normpath(path)
        path = path.replace('\\', '/')
        head, folder = os.path.split(path)
        new_dir_path = ''
        isdir = False
        while isdir == False:
            if not os.path.isdir(head):
                head, tail = os.path.split(head)
                new_dir_path = new_dir_path+tail + r"/"
            else:
                root = head
                isdir = True
        else:
            new_dir_path = folder+r"/"+new_dir_path
            new_dir_path = os.path.normpath(new_dir_path)
            head, tail = os.path.split(new_dir_path)
            temp = ''
            while tail:
                temp = temp + r"/" + tail
                dir_path = root+temp
                if not os.path.isdir(dir_path):
                    os.mkdir(dir_path)
                head, tail = os.path.split(head)

        return True
    except Exception as e:
        print(e)
        # logger.error('批量創建目錄出錯：{}'.format(e))
        return False


def mkdirs_in_batch_bak(path):
    try:
        path = os.path.normpath(path)
        path = path.replace('\\', '/')
        head, tail = os.path.split(path)
        if not os.path.isdir(path) and os.path.isfile(path):
            head, tail = os.path.split(head)

        if tail == '':
            return True

        new_dir_path = ''
        root = ''
        while tail:
            new_dir_path = new_dir_path + tail + '/'
            head, tail = os.path.split(head)
            root = head
        else:
            new_dir_path = root + new_dir_path
            new_dir_path = os.path.normpath(new_dir_path)
            head, tail = os.path.split(new_dir_path)
            temp = ''
            while tail:
                temp = temp + '/' + tail
                dir_path = root + temp
                if not os.path.isdir(dir_path):
                    os.mkdir(dir_path)
                head, tail = os.path.split(head)
        return True

    except Exception as e:
        # logger.error('批量創建目錄出錯：%s' % e)
        return False


@ api_view(['POST'])
@ permission_classes([AllowAny])
@authentication_classes([JWTAuthentication])
def outout_excel(request):
    result = {}
    print(request.data)
    try:
        dateFile = Mylist.objects.filter(complete=True,
                                         date__range=[request.data.get('startDate'), request.data.get('endDate')])

        wb = openpyxl.Workbook()
        sheet = wb.create_sheet("data", 0)
        sheet.cell(row=1, column=1).value = 'ID'
        sheet.cell(row=1, column=2).value = 'file'
        sheet.cell(row=1, column=3).value = 'date'
        sheet.cell(row=1, column=4).value = 'done_User'
        sheet.cell(row=1, column=5).value = 'starTime'
        sheet.cell(row=1, column=6).value = 'endTime'

        i = 2
        for j in dateFile:
            sheet.cell(row=i, column=1).value = j.name
            sheet.cell(row=i, column=2).value = j.file
            sheet.cell(row=i, column=3).value = j.date
            sheet.cell(row=i, column=4).value = j.attachmentfile.done_user
            sheet.cell(row=i, column=5).value = j.attachmentfile.start_time
            sheet.cell(row=i, column=6).value = j.attachmentfile.done_time
            i += 1

        fileName = "{}T{}".format(request.data.get('starTime'),
                                  request.data.get('endTime'))
        file_path = settings.MEDIA_ROOT+r"/" + \
            "excel" + r"/"+"{}.xlsx".format(fileName)
        wb.save(file_path)
        wb.close()

        # result['msg'] = 'Update Success'
        # result['success'] = True
        # return Response(result, status=status.HTTP_200_OK)

        file = open(file_path, 'rb')
        file_response = FileResponse(file)
        file_response['Content-Type'] = 'application/octet-stream'
        file_response["Access-Control-Expose-Headers"] = 'Content-Disposition'
        file_response['Content-Disposition'] = 'attachment;filename={}'.format(
            urlquote("{}.xlsx".format(fileName)))
        return file_response

    except Exception as e:
        result['msg'] = '{}'.format(e)
        result['success'] = False
        return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)
